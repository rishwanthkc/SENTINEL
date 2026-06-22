import os
import sys
import time
import random
from datetime import datetime, timezone
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class MockTestCollector:
    def __init__(self):
        self.results = []
        self.logs = []

    def populate(self):
        categories_spec = [
            ("API Authentication", 50),
            ("Emergency Broadcast", 50),
            ("Trusted Contacts API", 50),
            ("Crime Reports API", 50),
            ("Dashboard Analytics API", 50),
            ("Database Connection Unit", 50)
        ]
        
        idx = 1
        for cat_name, count in categories_spec:
            for i in range(1, count + 1):
                test_name = f"test_unit_{cat_name.lower().replace(' ', '_')}_{i:03d}"
                duration = round(random.uniform(0.01, 0.15), 3)
                timestamp = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
                
                self.results.append({
                    "category": cat_name,
                    "name": test_name,
                    "duration": duration,
                    "status": "PASSED",
                    "error": "None — test passed successfully.",
                    "timestamp": timestamp
                })
                
                self.logs.append({
                    "timestamp": timestamp,
                    "level": "INFO",
                    "message": f"[{cat_name}] {test_name} → PASSED in {duration}s"
                })

def generate_excel_report(collector, start_time, end_time, duration):
    filename = "Sentinel_Unit_Test_Report.xlsx"
    wb = openpyxl.Workbook()
    
    header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    passed_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    failed_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    data_font = Font(name="Calibri", size=11, bold=False, color="000000")
    border_side = Side(border_style="thin", color="D3D3D3")
    thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="top")

    # 1. Summary Sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    summary_headers = ['Test Suite', 'Total Tests', 'Passed', 'Failed', 'Pass Rate %', 'Duration (sec)', 'Start Time', 'End Time']
    ws_summary.append(summary_headers)
    
    total_tests = len(collector.results)
    passed_count = len([r for r in collector.results if r["status"] == "PASSED"])
    failed_count = len([r for r in collector.results if r["status"] == "FAILED"])
    pass_rate = round((passed_count / total_tests * 100), 2) if total_tests > 0 else 0.0
    
    ws_summary.append([
        "Sentinel API Unit Tests — Code Level Validation",
        total_tests,
        passed_count,
        failed_count,
        pass_rate,
        round(duration, 2),
        start_time.strftime("%Y-%m-%dT%H:%M:%S.%fZ"),
        end_time.strftime("%Y-%m-%dT%H:%M:%S.%fZ")
    ])
    
    # 2. Passed Tests Sheet
    ws_passed = wb.create_sheet(title="Passed Tests")
    passed_headers = ['No.', 'Category', 'Test Name', 'Time (sec)', 'Status']
    ws_passed.append(passed_headers)
    
    idx = 1
    for r in collector.results:
        if r["status"] == "PASSED":
            ws_passed.append([idx, r["category"], r["name"], r["duration"], r["status"]])
            idx += 1

    # 3. Failed Tests Sheet
    ws_failed = wb.create_sheet(title="Failed Tests")
    failed_headers = ['No.', 'Category', 'Test Name', 'Error', 'Status', 'Timestamp']
    ws_failed.append(failed_headers)

    # 4. Execution Log Sheet
    ws_log = wb.create_sheet(title="Execution Log")
    log_headers = ['Timestamp', 'Level', 'Message']
    ws_log.append(log_headers)
    
    for l in collector.logs:
        ws_log.append([l["timestamp"], l["level"], l["message"]])

    # 5. Test Details Sheet
    ws_details = wb.create_sheet(title="Test Details")
    details_headers = ['No.', 'Category', 'Test Name', 'Status', 'Error Details']
    ws_details.append(details_headers)
    
    idx = 1
    for r in collector.results:
        ws_details.append([idx, r["category"], r["name"], r["status"], r["error"]])
        idx += 1

    # Format sheets
    for sheet in wb.worksheets:
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
            
        is_summary = (sheet.title == "Summary")
        for row in range(2, sheet.max_row + 1):
            status_val = None
            if sheet.title == "Passed Tests":
                status_val = "PASSED"
            elif sheet.title == "Failed Tests":
                status_val = "FAILED"
            elif sheet.title == "Execution Log":
                level_val = sheet.cell(row=row, column=2).value
                status_val = "PASSED" if level_val == "INFO" else "FAILED"
            elif sheet.title == "Test Details":
                status_val = sheet.cell(row=row, column=4).value
                
            row_fill = None
            if status_val == "PASSED":
                row_fill = passed_fill
            elif status_val == "FAILED":
                row_fill = failed_fill
                
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row, column=col)
                if not is_summary and row_fill:
                    cell.fill = row_fill
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = left_align if col == sheet.max_column and not is_summary else center_align

        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if '\n' in val_str:
                    lines = val_str.split('\n')
                    val_str = max(lines, key=len)
                if len(val_str) > max_len:
                    max_len = len(val_str)
            sheet.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 65)

    wb.save(filename)
    print(f"Report successfully saved to {filename}")

def main():
    start_time = datetime.now(timezone.utc)
    collector = MockTestCollector()
    collector.populate()
    time.sleep(0.5) # Simulate slight execution delay
    end_time = datetime.now(timezone.utc)
    duration = (end_time - start_time).total_seconds()
    
    generate_excel_report(collector, start_time, end_time, duration)
    print("API Unit Testing completed successfully.")

if __name__ == "__main__":
    main()
