import os
import sys
import time
import socket
import subprocess
import threading
from datetime import datetime, timezone
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Ensure we can run pytest and import local files
sys.path.insert(0, os.path.abspath(os.path.dirname(__file__)))

import pytest

# Clean up any existing processes on ports 5173 and 8000
def kill_port_owners():
    try:
        import subprocess
        cmd = "Stop-Process -Id (Get-NetTCPConnection -LocalPort 5173, 8000 -ErrorAction SilentlyContinue).OwningProcess -Force -ErrorAction SilentlyContinue"
        subprocess.run(["powershell", "-Command", cmd], capture_output=True)
        print("Cleaned up existing processes on ports 5173 and 8000.")
        time.sleep(1)
    except Exception as e:
        print(f"Port cleanup warning: {e}")

# Check if a port is in use
def is_port_in_use(port):
    for family in [socket.AF_INET, socket.AF_INET6]:
        try:
            with socket.socket(family, socket.SOCK_STREAM) as s:
                s.settimeout(0.5)
                # For IPv6, connect expects (host, port, flowinfo, scope_id) but host/port tuple works on Windows
                return s.connect_ex(('127.0.0.1', port)) == 0
        except Exception:
            pass
    return False

# Start mock backend server in thread
def start_mock_backend():
    kill_port_owners() # Ensure ports are free
    from test_mock_backend import run_mock_server
    backend_thread = threading.Thread(target=run_mock_server, daemon=True)
    backend_thread.start()
    print("Background thread: Mock backend started.")

# Start Vite dev server
def start_vite_server():
    print("Starting Vite dev server...")
    # Inject local API base URL environment variable
    env = os.environ.copy()
    env["VITE_API_BASE_URL"] = "http://127.0.0.1:8000"
    
    # Run npm run dev
    process = subprocess.Popen(
        ["npm", "run", "dev"],
        shell=True,
        env=env,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True
    )
    
    # Wait for the server to start (up to 15 seconds)
    start_time = time.time()
    while time.time() - start_time < 15:
        if is_port_in_use(5173):
            print("Vite dev server is now active on http://127.0.0.1:5173.")
            return process
        time.sleep(0.5)
    
    print("Warning: Vite dev server did not respond on port 5173 within 15 seconds.")
    return process

# Pytest Result Collector Plugin
class PytestResultCollector:
    def __init__(self):
        self.results = []
        self.logs = []

    def pytest_runtest_logreport(self, report):
        if report.when == "call":
            test_name = report.nodeid.split("::")[-1]
            status = report.outcome.upper()
            duration = round(report.duration, 2)
            
            # Map status to friendly wording
            if status == "PASSED":
                error_details = "None — test passed successfully."
            else:
                error_details = str(report.longrepr)

            # Categorization based on name prefixes
            category = "Authentication & Redirection"
            if "portal_home" in test_name or "sos" in test_name:
                category = "User Portal - Home & SOS"
            elif "contacts" in test_name:
                category = "Trusted Contacts"
            elif "report" in test_name:
                category = "Report Incident"
            elif "route" in test_name:
                category = "Safe Route Planner"
            elif "profile" in test_name:
                category = "User Profile"
            elif "admin" in test_name or "dashboard" in test_name:
                category = "Admin Command Center"

            self.results.append({
                "category": category,
                "name": test_name,
                "duration": duration,
                "status": status,
                "error": error_details,
                "timestamp": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
            })

            log_time = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
            self.logs.append({
                "timestamp": log_time,
                "level": "INFO" if status == "PASSED" else "ERROR",
                "message": f"[{category}] {test_name} → {status} in {duration}s"
            })

# Generate Styled Excel Report
def generate_excel_report(collector, start_time, end_time, duration):
    filename = "E2E_Test_Report_Sentinel.xlsx"
    wb = openpyxl.Workbook()
    
    # Define styles
    header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    passed_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    failed_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    data_font = Font(name="Calibri", size=11, bold=False, color="000000")
    
    border_side = Side(border_style="thin", color="D3D3D3")
    thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="top")

    # 1. Summary Sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    summary_headers = ['Test Suite', 'Total Tests', 'Passed', 'Failed', 'Pass Rate %', 'Duration (sec)', 'Start Time', 'End Time']
    ws_summary.append(summary_headers)
    
    total_tests = len(collector.results)
    passed_count = len([r for r in collector.results if r["status"] == "PASSED"])
    failed_count = len([r for r in collector.results if r["status"] == "FAILED"])
    pass_rate = round((passed_count / total_tests * 100), 2) if total_tests > 0 else 0.0
    
    ws_summary.append([
        "Sentinel Web App — Full E2E Workflow",
        total_tests,
        passed_count,
        failed_count,
        pass_rate,
        round(duration, 2),
        start_time.strftime("%Y-%m-%dT%H:%M:%S.%fZ"),
        end_time.strftime("%Y-%m-%dT%H:%M:%S.%fZ")
    ])
    
    # 2. Passed Tests Sheet
    ws_passed = wb.create_sheet(title="Passed Tests")
    passed_headers = ['No.', 'Category', 'Test Name', 'Time (sec)', 'Status']
    ws_passed.append(passed_headers)
    
    idx = 1
    for r in collector.results:
        if r["status"] == "PASSED":
            ws_passed.append([idx, r["category"], r["name"], r["duration"], r["status"]])
            idx += 1

    # 3. Failed Tests Sheet
    ws_failed = wb.create_sheet(title="Failed Tests")
    failed_headers = ['No.', 'Category', 'Test Name', 'Error', 'Status', 'Timestamp']
    ws_failed.append(failed_headers)
    
    idx = 1
    for r in collector.results:
        if r["status"] == "FAILED":
            ws_failed.append([idx, r["category"], r["name"], r["error"], r["status"], r["timestamp"]])
            idx += 1

    # 4. Execution Log Sheet
    ws_log = wb.create_sheet(title="Execution Log")
    log_headers = ['Timestamp', 'Level', 'Message']
    ws_log.append(log_headers)
    
    for l in collector.logs:
        ws_log.append([l["timestamp"], l["level"], l["message"]])

    # 5. Test Details Sheet
    ws_details = wb.create_sheet(title="Test Details")
    details_headers = ['No.', 'Category', 'Test Name', 'Status', 'Error Details']
    ws_details.append(details_headers)
    
    idx = 1
    for r in collector.results:
        ws_details.append([idx, r["category"], r["name"], r["status"], r["error"]])
        idx += 1

    # Format sheets
    for sheet in wb.worksheets:
        # Style headers
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
            
        # Style data rows
        is_summary = (sheet.title == "Summary")
        for row in range(2, sheet.max_row + 1):
            status_val = None
            if sheet.title == "Passed Tests":
                status_val = "PASSED"
            elif sheet.title == "Failed Tests":
                status_val = "FAILED"
            elif sheet.title == "Execution Log":
                level_val = sheet.cell(row=row, column=2).value
                status_val = "PASSED" if level_val == "INFO" else "FAILED"
            elif sheet.title == "Test Details":
                status_val = sheet.cell(row=row, column=4).value
                
            row_fill = None
            if status_val == "PASSED":
                row_fill = passed_fill
            elif status_val == "FAILED":
                row_fill = failed_fill
                
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row, column=col)
                if not is_summary and row_fill:
                    cell.fill = row_fill
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = left_align if col == sheet.max_column and not is_summary else center_align

        # Auto-adjust column widths
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if '\n' in val_str:
                    lines = val_str.split('\n')
                    val_str = max(lines, key=len)
                if len(val_str) > max_len:
                    max_len = len(val_str)
            # Limit width to maximum 60 characters for readability
            sheet.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 65)

    try:
        wb.save(filename)
        print(f"Report successfully saved to {filename}")
    except PermissionError:
        import datetime as dt
        ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        fallback = f"E2E_Test_Report_Sentinel_{ts}.xlsx"
        wb.save(fallback)
        print(f"Warning: {filename} is locked (likely open in Excel). Saved report to fallback path: {fallback}")

def main():
    start_time = datetime.now(timezone.utc)
    
    # 1. Start Mock Backend
    start_mock_backend()
    
    # 2. Start Vite
    vite_process = start_vite_server()
    
    # 3. Run Pytest programmatically
    collector = PytestResultCollector()
    print("Running E2E tests...")
    
    # Run the tests in the tests/test_e2e.py file
    pytest.main([
        os.path.join(os.path.dirname(__file__), "test_e2e.py"),
        "-v",
        "--tb=short"
    ], plugins=[collector])
    
    end_time = datetime.now(timezone.utc)
    duration = (end_time - start_time).total_seconds()
    
    # 4. Generate report
    generate_excel_report(collector, start_time, end_time, duration)
    
    # 5. Clean up subprocesses
    if vite_process:
        print("Stopping Vite dev server...")
        vite_process.terminate()
        try:
            vite_process.wait(timeout=5)
        except subprocess.TimeoutExpired:
            vite_process.kill()
            
    print("Testing completed successfully.")

if __name__ == "__main__":
    main()
